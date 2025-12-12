import openpyxl as op
from openpyxl.styles import Alignment

class ExcelSet():
    ### 인스턴스 변수
    def __init__(self):
        self.wb = None
        self.ws = None
    

    ### 엑셀 파일 관련 기능
    # 파일 생성
    def create_file(self):
        self.wb = op.Workbook()
        
    # 파일 저장
    def save_file(self, name):
        self.wb.save(name)

    # 파일 닫기
    def close_file(self):
        self.wb.close()
    

    ### 엑셀 sheet 관련 기능
    # sheet 생성
    def create_sheet(self, sheet_name):
        self.wb.create_sheet(sheet_name)

    # 기본 시트 이름 변경
    def change_default_sheet_name(self, sheet_name, default_sheet_name):
        self.ws = self.wb[default_sheet_name]
        self.ws.title = sheet_name       

    
    ### 행 탐색 관련 기능
    # 행(데이터 존재) 위치 찾기
    def find_last_data_row(self):
        current_row = self.ws.max_row
        return current_row
    
    ### 데이터 작성 관련 기능
    # 열 이름 작성
    def write_column_name(self, index, data):
        self.ws.cell(row = 1, column = index + 1).value = data[index]

    def write_data(self, index, data, jndex):
        self.ws.cell(row = index + 2, column = jndex + 1).value = data[index][jndex]
